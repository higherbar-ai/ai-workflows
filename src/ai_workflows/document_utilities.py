#  Copyright (c) 2024 Higher Bar AI, PBC
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

"""Utilities for reading and processing documents for AI workflows."""

from ai_workflows.llm_utilities import LLMInterface
import fitz  # (PyMuPDF)
from fitz.utils import get_pixmap
import pymupdf4llm
from PIL import Image
import io
import base64
from langchain_core.messages import HumanMessage
import json
import subprocess
import os
from typing import Tuple, List, Union, Dict, Optional, Any, Callable
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell, Cell
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.merge import MergedCellRange
from unstructured.partition.auto import partition
from unstructured.documents.elements import (
    Text, Title, NarrativeText, ListItem, Table, Image as ImageElement, PageBreak,
    Header, Footer, Address
)
from dataclasses import dataclass
from pathlib import Path
import re
import tempfile
import logging
import tiktoken


class DocumentInterface:
    """Utility class for reading and processing documents for AI workflows."""

    # class-level member variables
    llm_interface: LLMInterface = None
    via_pdf_file_extensions = [".docx", ".doc", ".pptx"]
    pymupdf_file_extensions = [".pdf", ".xps", ".epub", ".mobi", ".fb2", ".cbz", ".svg", ".txt"]
    libreoffice_file_extensions = [
        ".odt", ".csv", ".db", ".doc", ".docx", ".dotx", ".fodp", ".fods", ".fodt", ".mml", ".odb", ".odf", ".odg",
        ".odm", ".odp", ".ods", ".otg", ".otp", ".ots", ".ott", ".oxt", ".pdf", ".pptx", ".psw", ".sda", ".sdc", ".sdd",
        ".sdp", ".sdw", ".slk", ".smf", ".stc", ".std", ".sti", ".stw", ".sxc", ".sxg", ".sxi", ".sxm", ".sxw", ".uof",
        ".uop", ".uos", ".uot", ".vsd", ".vsdx", ".wdb", ".wps", ".wri", ".xls", ".xlsx"
    ]
    max_xlsx_via_pdf_pages: int = 10
    max_json_via_markdown_pages = 50
    max_json_via_markdown_tokens = 25000

    def __init__(self, llm_interface: LLMInterface = None):
        """
        Initialize the document interface for reading and processing documents.

        :param llm_interface: LLM interface for interacting with LLMs in AI workflows (defaults to None, which won't
            use an LLM to convert supported document types to markdown).
        :type llm_interface: LLMInterface
        """

        # if specified, remember our LLM interface
        if llm_interface:
            self.llm_interface = llm_interface

    def convert_to_markdown(self, filepath: str) -> str:
        """
        Convert a document to markdown.

        :param filepath: Path to the file.
        :type filepath: str
        :return: Markdown output.
        :rtype: str
        """

        # use internal conversion function
        return self._convert(filepath, to_format="md")

    def convert_to_json(self, filepath: str, json_context: str, json_job: str, json_output_spec: str,
                        markdown_first: Optional[bool] = None) -> list[dict]:
        """
        Convert a document to JSON.

        :param filepath: Path to the file.
        :type filepath: str
        :param json_context: Context for the LLM prompt used in JSON conversion (e.g., "The file contains a survey
          instrument administered by trained enumerators to households in Zimbabwe."). (Required for JSON output.)
        :type json_context: str
        :param json_job: Description of the job to do for the LLM prompt used in JSON conversion (e.g., "Your job is to
          extract each question or form field included in the text or page given."). (Required for JSON output.)
        :type json_job: str
        :param json_output_spec: JSON output specification for the LLM prompt (e.g., "Respond in correctly-formatted
          JSON with a single key named `questions` that is a list of dicts, one for each question or form field, each
          with the keys listed below..."). (Required for JSON output.)
        :type json_output_spec: str
        :param markdown_first: Whether to convert to Markdown first and then to JSON using an LLM. Set this to true if
          page-by-page conversion is not working well for elements that span pages; the Markdown-first approach will
          convert page-by-page to Markdown and then convert to JSON as the next step. When going the Markdown route,
          the JSON conversion will take place in a single step, with all of the Markdown supplied at once — so it
          won't work if the file Markdown is too large to fit in the LLM context window. The default is None, which will
          use the Markdown path for smaller PDF files and the page-by-page path for larger ones.
        :type markdown_first: Optional[bool]
        :return: List of dicts from page-level JSON results.
        :rtype: list[dict]
        """

        if markdown_first is None:
            # figure out whether we should convert to Markdown page-by-page and then JSON all in one go or convert
            # to JSON page-by-page directly

            # extract file extension
            ext = os.path.splitext(filepath)[1].lower()

            # if we're going to convert from PDF using an LLM, we need to figure out the right choice
            if self.llm_interface and (ext == '.pdf' or ext in DocumentInterface.via_pdf_file_extensions):
                # convert to Markdown without LLM assistance
                doc_interface_no_llm = DocumentInterface()
                markdown = doc_interface_no_llm.convert_to_markdown(filepath)

                # if PDF doesn't have much text, it might be scanned or image-based and require OCR
                if len(markdown) < 200 and ext == '.pdf':
                    # check number of PDF pages and decide whether to use Markdown or direct-to-JSON
                    doc = fitz.open(filepath)
                    if len(doc) <= DocumentInterface.max_json_via_markdown_pages:
                        # if we're within the limit, use Markdown conversion first
                        markdown_first = True
                    else:
                        # if we're over the limit, use page-by-page JSON conversion
                        markdown_first = False
                else:
                    # use tokens to decide
                    encoding = tiktoken.encoding_for_model(self.llm_interface.model)
                    if len(encoding.encode(markdown)) <= DocumentInterface.max_json_via_markdown_tokens:
                        # if we're within the limit, use Markdown conversion first
                        markdown_first = True
                    else:
                        # if we're over the limit, use page-by-page JSON conversion
                        markdown_first = False
            else:
                # for other file types or without an LLM, they always use Markdown first anyway
                markdown_first = True

        # use internal conversion function
        return self._convert(filepath, to_format="json" if not markdown_first else "mdjson",
                             json_context=json_context, json_job=json_job, json_output_spec=json_output_spec)

    def _convert(self, filepath: str, to_format: str = "md", json_context: str = "", json_job: str = "",
                 json_output_spec: str = "") -> str | list[dict]:
        """
        Convert a document to Markdown or JSON.

        :param filepath: Path to the file.
        :type filepath: str
        :param to_format: Format to convert to ("md" for Markdown, "json" for JSON, or "mdjson" for JSON from Markdown).
          Default is "md" for Markdown. The "mdjson" option is a special case that converts to Markdown first and then
          to JSON using an LLM. In contrast, the "json" option converts directly to JSON using an LLM when it can,
          bypassing the Markdown step (but when it does, it processes the document page-by-page, which can lead to
          worse results if elements span pages). Note that all JSON conversion requires an LLM interface be passed to
          the DocumentInterface constructor.
        :type to_format: str
        :param json_context: Context for the LLM prompt used in JSON conversion (e.g., "The file contains a survey
          instrument administered by trained enumerators to households in Zimbabwe."). (Required for JSON output.)
        :type json_context: str
        :param json_job: Description of the job to do for the LLM prompt used in JSON conversion (e.g., "Your job is to
          extract each question or form field included in the text or page given."). (Required for JSON output.)
        :type json_job: str
        :param json_output_spec: JSON output specification for the LLM prompt (e.g., "Respond in correctly-formatted
          JSON with a single key named `questions` that is a list of dicts, one for each question or form field, each
          with the keys listed below..."). (Required for JSON output.)
        :type json_output_spec: str
        :return: Markdown output or list of dicts containing JSON results.
        :rtype: str | list[dict]
        """

        # validate parameters
        if to_format not in ["md", "json", "mdjson"]:
            raise ValueError("Invalid 'to_format' parameter; must be 'md' for Markdown, 'json' for JSON, or 'mdjson' "
                             "for JSON via Markdown.")
        if to_format in ["json", "mdjson"] and (not json_context or not json_job or not json_output_spec):
            raise ValueError("For JSON output, 'context', 'job_to_do', and 'output_format' parameters are required.")
        if to_format in ["json", "mdjson"] and self.llm_interface is None:
            raise ValueError("LLM interface required for JSON output.")

        # extract file extension
        ext = os.path.splitext(filepath)[1].lower()

        # if we have an LLM interface, use it when we can
        if self.llm_interface is not None:
            # always convert PDFs with the LLM
            if ext == '.pdf':
                # convert PDF using LLM
                pdf_converter = PDFDocumentConverter(self.llm_interface)
                if to_format == "md":
                    # convert to Markdown
                    return pdf_converter.pdf_to_markdown(filepath)
                elif to_format == "json":
                    # convert directly to JSON
                    return pdf_converter.pdf_to_json(filepath, json_context, json_job, json_output_spec)
                else:
                    # convert to Markdown and then to JSON
                    markdown = pdf_converter.pdf_to_markdown(filepath)
                    return self.markdown_to_json(markdown, json_context, json_job, json_output_spec)

            # convert certain other file types to PDF to then convert with the LLM
            if ext in DocumentInterface.via_pdf_file_extensions:
                with tempfile.TemporaryDirectory() as temp_dir:
                    # convert to PDF in temporary directory
                    pdf_path = self.convert_to_pdf(filepath, temp_dir)
                    # convert PDF using LLM
                    pdf_converter = PDFDocumentConverter(self.llm_interface)
                    if to_format == "md":
                        # convert to Markdown
                        return pdf_converter.pdf_to_markdown(pdf_path)
                    elif to_format == "json":
                        # convert directly to JSON
                        return pdf_converter.pdf_to_json(pdf_path, json_context, json_job, json_output_spec)
                    else:
                        # convert to Markdown and then to JSON
                        markdown = pdf_converter.pdf_to_markdown(pdf_path)
                        return self.markdown_to_json(markdown, json_context, json_job, json_output_spec)

        # if Excel, see if we can convert to Markdown using our custom converter
        if ext == '.xlsx':
            # convert Excel to Markdown using custom converter
            # (try to keep images and charts if we have an LLM available and we're after Markdown output)
            result, markdown = (ExcelDocumentConverter.convert_excel_to_markdown
                                (filepath, lose_unsupported_content=(not (self.llm_interface and to_format == "md"))))
            if result:
                if to_format == "json":
                    # if we're after JSON, convert the Markdown to JSON using the LLM
                    return self.markdown_to_json(markdown, json_context, json_job, json_output_spec)
                else:
                    # otherwise, just return the Markdown
                    return markdown
            else:
                # log reason from returned Markdown
                logging.info(f"Failed to convert {filepath} to Markdown: {markdown}")

                # if we have an LLM and we're after Markdown, PDF it and then convert with LLM if we can
                # (we don't want to use an LLM on Excel files headed for JSON)
                if self.llm_interface is not None and to_format == "md":
                    with (tempfile.TemporaryDirectory() as temp_dir):
                        # convert to PDF in temporary directory
                        pdf_path = self.convert_to_pdf(filepath, temp_dir)

                        # check number of PDF pages and only move forward with LLM conversion if it's within the limit
                        doc = fitz.open(pdf_path)
                        if len(doc) <= DocumentInterface.max_xlsx_via_pdf_pages:
                            # convert PDF to Markdown using LLM
                            pdf_converter = PDFDocumentConverter(self.llm_interface)
                            return pdf_converter.pdf_to_markdown(pdf_path)
                        else:
                            logging.info(f"{filepath} converted to {len(doc)} pages, which is over the limit "
                                         f"({DocumentInterface.max_xlsx_via_pdf_pages}); converting without images or "
                                         f"charts...")
                            result, markdown = (ExcelDocumentConverter.convert_excel_to_markdown
                                                (filepath, lose_unsupported_content=True))
                            if result:
                                return markdown
                            else:
                                # log reason from returned Markdown
                                # then fall through to let Unstructured have a try at it
                                logging.info(f"Failed to convert {filepath} to Markdown: {markdown}")

        # otherwise, see if we can use PyMuPDFLLM to convert (includes .pdf, .txt, .svg, .cbz, .epub, .mobi, .xps)
        if ext in DocumentInterface.pymupdf_file_extensions:
            markdown = pymupdf4llm.to_markdown(filepath)
        else:
            # otherwise, use Unstructured to convert
            doc_converter = UnstructuredDocumentConverter()
            markdown = doc_converter.convert_to_markdown(filepath)

        if to_format in ["json", "mdjson"]:
            # if we're after JSON, convert the Markdown to JSON using the LLM
            return self.markdown_to_json(markdown, json_context, json_job, json_output_spec)
        else:
            # otherwise, just return the Markdown
            return markdown

    @staticmethod
    def convert_to_pdf(filepath: str, output_dir: str) -> str:
        """
        Convert a document to PDF using LibreOffice.

        :param filepath: Path to the document file.
        :type filepath: str
        :param output_dir: Path to the output directory.
        :type output_dir: str
        :return: Path to the converted PDF file. Throws exception on failure.
        :rtype: str
        """

        # confirm that it's a format we can convert (i.e., a supported LibreOffice format)
        ext = os.path.splitext(filepath)[1].lower()
        if ext not in DocumentInterface.libreoffice_file_extensions:
            raise ValueError(f"Unsupported file format: {ext}. convert_to_pdf() only support formats that LibreOffice "
                             f"supports.")

        # call to LibreOffice to convert the document to PDF format
        subprocess.run([
            'soffice', '--headless', '--convert-to', 'pdf',
            filepath, '--outdir', output_dir
        ], check=True)

        # return path to the converted PDF file
        return os.path.join(output_dir, os.path.splitext(os.path.basename(filepath))[0] + '.pdf')

    def markdown_to_json(self, markdown: str, json_context: str, json_job: str, json_output_spec: str) -> list[dict]:
        """
        Convert Markdown text to JSON using an LLM.

        :param markdown: Markdown text to convert to JSON.
        :type markdown: str
        :param json_context: Context for the LLM prompt (e.g., "The file contains a survey instrument administered by
          trained enumerators to households in Zimbabwe.").
        :type json_context: str
        :param json_job: Job to do for the LLM prompt (e.g., "Your job is to extract each question or form field
          included in the text.").
        :type json_job: str
        :param json_output_spec: Output format for the LLM prompt (e.g., "Respond in correctly-formatted JSON with a
          single key named `questions` that is a list of dicts, one for each question or form field, each with the keys
          listed below...").
        :type json_output_spec: str
        :return: List of dicts with JSON results (could be all results in one dict, could be split into multiple because
          of page-by-page or other batched processing).
        :rtype: list[dict]
        """

        # require LLM interface to continue
        if self.llm_interface is None:
            raise ValueError("LLM interface required for JSON conversion")

        # set up for JSON processing
        json_prompt = f"""Consider the Markdown text below, which has been extracted from a file.

{json_context}

{json_job}

{json_output_spec}

Markdown text enclosed by |@| delimiters:

|@|{markdown}|@|

Your JSON response precisely following the instructions given above the Markdown text:"""

        # for now, process all in one go (assumes it all fits in the LLM context window)
        response_text, response_dict = self.llm_interface.process_json_response(
            self.llm_interface.llm_json_response_with_timeout(
                [HumanMessage(content=[{"type": "text", "text": json_prompt}])]))

        # raise exception if we didn't get a response
        if response_dict is None:
            logging.error(f"ERROR: Error extracting JSON from Markdown: {response_text}")
            raise ValueError(f"Error extracting JSON from Markdown: {response_text}")

        # log all returned elements
        logging.info(f"Extracted JSON from Markdown: {json.dumps(response_dict, indent=2)}")

        # return results
        return [response_dict]


class PDFDocumentConverter:
    """Utility class for converting PDF files to Markdown."""

    # class-level member variables
    llm_interface: LLMInterface = None

    def __init__(self, llm_interface: LLMInterface = None):
        """
        Initialize for converting PDF files.

        :param llm_interface: LLM interface for interacting with LLMs in AI workflows (defaults to None, which won't
            use an LLM to convert PDF files to Markdown).
        :type llm_interface: LLMInterface
        """

        # if specified, remember our LLM interface
        if llm_interface:
            self.llm_interface = llm_interface

    @staticmethod
    def pdf_to_images(pdf_path: str, dpi: int = 300) -> list[Image.Image]:
        """
        Convert a PDF to a list of PIL Images.

        This function opens a PDF file, renders each page as an image at the specified DPI, and returns a list of these
        images.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :param dpi: DPI to use for rendering the PDF. Default is 300.
        :type dpi: int
        :return: List of PIL Images representing the pages within the PDF.
        :rtype: list[Image.Image]
        """

        # open PDF
        doc = fitz.open(pdf_path)

        # calculate scaling factor based on DPI
        zoom = dpi / 72  # 72 is the default PDF DPI
        matrix = fitz.Matrix(zoom, zoom)

        # first pass to determine minimum page dimensions
        min_page_width = None
        min_page_height = None
        for page in doc:
            page_width = page.rect.width
            page_height = page.rect.height

            if min_page_width is None or page_width < min_page_width:
                min_page_width = page_width
            if min_page_height is None or page_height < min_page_height:
                min_page_height = page_height

        # tolerance for dimension comparison
        tolerance = 1e-2  # Adjust if necessary for your PDFs

        # function to check if two dimensions are approximately equal
        def approx_equal(a, b, tol=tolerance):
            return abs(a - b) / max(abs(a), abs(b)) < tol

        # second pass to process pages
        images = []
        for page in doc:
            # get page dimensions
            page_width = page.rect.width
            page_height = page.rect.height

            # render page to pixmap
            pix = get_pixmap(page, matrix=matrix, alpha=False)

            # convert to PIL Image
            img = Image.frombytes("RGB", (pix.w, pix.h), pix.samples)

            # check if page is a double-page spread
            if approx_equal(page_height, min_page_height) and approx_equal(page_width, 2 * min_page_width):
                # if so, split the image into two images
                left_img = img.crop((0, 0, img.width // 2, img.height))
                right_img = img.crop((img.width // 2, 0, img.width, img.height))
                images.extend([left_img, right_img])
            else:
                images.append(img)

        doc.close()
        return images

    @staticmethod
    def _get_image_bytes(image: Image.Image, output_format: str = 'PNG') -> bytes:
        """
        Convert a PIL Image to bytes in the specified format.

        This function takes a PIL Image object and converts it to a byte array in the specified format.

        :param image: PIL Image to convert.
        :type image: Image.Image
        :param output_format: Output format for the image (default is 'PNG').
        :type output_format: str
        :return: Bytes representing the image in the specified format.
        :rtype: bytes
        """

        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format=output_format)
        return img_byte_arr.getvalue()

    @staticmethod
    def _starts_with_heading(content: str) -> bool:
        """
        Check if the content appears to start with a heading.

        This function checks if the given content starts with a heading, either in ATX or Setext style.

        :param content: Content to check.
        :type content: str
        :return: True if the content starts with a heading, False otherwise.
        :rtype: bool
        """

        # split content into lines
        lines = content.splitlines()
        i = 0
        # skip blank lines
        while i < len(lines) and lines[i].strip() == '':
            i += 1
        if i >= len(lines):
            # if no non-blank lines, return False
            return False
        # get first non-blank line, stripping whitespace
        line = lines[i].strip()
        # check for ATX heading (starts with '#')
        if line.startswith('#'):
            # we found a heading, so return True
            return True
        # check for setext-style heading (underlines with '=' or '-')
        if i + 1 < len(lines):
            underline = lines[i + 1].strip()
            if len(underline) >= len(line) and (all(c == '-' for c in underline) or all(c == '=' for c in underline)):
                # we found a heading, so return True
                return True

        # if no heading found, return False
        return False

    @staticmethod
    def _clean_and_reorder_elements(elements: list[dict]) -> list[dict]:
        """
        Clean and reorder elements, dropping page headers and footers and reordering body text sections as needed to
        ensure uninterrupted cross-page flow within sections.

        :param elements: List of elements to clean and reorder.
        :type elements: list[dict]
        :return: Cleaned and reordered list of elements.
        :rtype: list[dict]
        """

        output_elements = []
        last_body_text_idx = None
        for element in elements:
            element_type = element.get('type')
            # drop elements with type 'page_header' or 'page_footer'
            if element_type in ['page_header', 'page_footer']:
                continue
            elif element_type == 'body_text_section':
                content = element.get('content', '')
                if PDFDocumentConverter._starts_with_heading(content):
                    # sections with headings can appear in their existing order
                    output_elements.append(element)
                    last_body_text_idx = len(output_elements) - 1
                else:
                    if last_body_text_idx is not None:
                        # body sections without headings should move to just after the last body section, if any
                        last_body_text_idx += 1
                        output_elements.insert(last_body_text_idx, element)
                    else:
                        # otherwise, if no prior body section, leave it where it is
                        output_elements.append(element)
                        last_body_text_idx = len(output_elements) - 1
            else:
                # other sections can stay in their existing order
                output_elements.append(element)

        # return trimmed and reordered list of elements
        return output_elements

    @staticmethod
    def _assemble_markdown(elements: list[dict]) -> str:
        """
        Assemble a list of elements into a single markdown output.

        :param elements: List of elements to assemble.
        :type elements: list[dict]
        :return: Markdown output.
        :rtype: str
        """

        markdown_output = ''
        for element in elements:
            element_type = element.get('type')
            content = element.get('content', '')
            if element_type == 'body_text_section':
                # just add body content with paragraph separation
                # (could be smarter about this, only adding paragraph breaks where it looks appropriate — but it's
                # not clear that over-breaking will overly affect downstream uses)
                markdown_output += content.strip() + '\n\n'
            else:
                # add other types of elements with labels and more visual separation
                markdown_output += f'***\n\n{element_type.upper()}:\n\n{content.strip()}\n\n***\n\n'

        # return assembled markdown output with extra newlines at the end stripped out
        return markdown_output.strip()

    def pdf_to_json(self, pdf_path: str, json_context: str, json_job: str, json_output_spec: str) -> list[dict]:
        """
        Process a PDF file page-by-page to extract elements and output JSON text.

        This function reads a PDF file, converts pages to images, processes each image with an LLM, and assembles the
        returned elements into a single JSON output.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :param json_context: Context for the LLM prompt (e.g., "The file contains a survey instrument.").
        :type json_context: str
        :param json_job: Job to do for the LLM prompt (e.g., "Your job is to extract each question or form field
          included on the page."). In this case, the job will be to process each page, one at a time.
        :type json_job: str
        :param json_output_spec: Output format for the LLM prompt (e.g., "Respond in correctly-formatted JSON with a
          single key named `questions` that is a list of dicts, one for each question or form field, each with the keys
          listed below...").
        :type json_output_spec: str
        :return: List of parsed results from all pages, one per page, in order.
        :rtype: list[dict]
        """

        # require LLM interface to continue
        if self.llm_interface is None:
            raise ValueError("LLM interface required for PDF to JSON conversion")

        # convert PDF to images
        images = PDFDocumentConverter.pdf_to_images(pdf_path)

        # set up for image processing
        image_prompt = f"""Consider the attached image, which shows a single page (or, sometimes, two facing pages) from a PDF file.

{json_context}

{json_job}

{json_output_spec}

Your JSON response precisely following the instructions above:"""

        # process each page
        all_dicts = []
        logging.log(logging.INFO, f"Processing PDF {pdf_path} from {len(images)} images")
        for i, img in enumerate(images):
            logging.log(logging.INFO, f"Processing PDF page {i + 1}: Size={img.size}, Mode={img.mode}")

            # encode image contents for OpenAI
            encoded_image = base64.b64encode(PDFDocumentConverter._get_image_bytes(img)).decode('utf-8')
            # call out to the LLM and process the returned JSON
            response_text, response_dict = self.llm_interface.process_json_response(
                self.llm_interface.llm_json_response_with_timeout([
                    HumanMessage(content=[
                        {"type": "text", "text": image_prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{encoded_image}"}},
                    ])]))

            # assemble results
            if response_dict is not None:
                # add response to running list of parsed results
                all_dicts.append(response_dict)

                # log all returned elements
                logging.info(f"Extracted JSON for page {i + 1}: {json.dumps(response_dict, indent=2)}")
            else:
                logging.error(f"ERROR: Error extracting JSON from page {i+1}: {response_text}")
                raise ValueError(f"Error extracting JSON from page {i+1} of {pdf_path}: {response_text}")

        # return all results
        return all_dicts

    def pdf_to_markdown(self, pdf_path: str) -> str:
        """
        Process a PDF file to extract elements and output Markdown text.

        This function reads a PDF file, converts it to images, processes each image with an LLM, and assembles the
        returned elements into a single markdown output. If no LLM is available, the function falls back to PyMuPDFLLM
        for Markdown conversion.

        :param pdf_path: Path to the PDF file.
        :type pdf_path: str
        :return: Markdown text.
        :rtype: str
        """

        # check for LLM interface
        if self.llm_interface is None:
            # since no LLM interface, use PyMuPDFLLM to convert PDF to Markdown
            return pymupdf4llm.to_markdown(pdf_path)

        # otherwise, we'll use the LLM to process the PDF
        json_context = "The page might include a mix of text, tables, figures, charts, images, or other elements."

        json_job = f"""Your job is to:

First, scan the image to identify each distinct element in the image, where each element is a part of the page that can be handled separately from the other parts of the page. Elements include, for example:

   1. The main body text (if any), possibly separated into sections. This is the primary text of the page, which can begin on prior pages and/or continue on to future pages.

   2. Boxout text (if any). These might be sidebars, callout boxes, or other separated sub-sections that are self-contained within the page.

   3. Tables (if any). These might include a title, the table itself, and, possibly, end notes or captions.

   4. Charts or graphs (if any). These might include a title, the chart or graph itself, and, possibly, notes or captions just beneath the chart or graph.

   5. Image or figure (if any). These might include a title, the image or figure itself, and, possibly, notes or captions just beneath the image or figure.

   6. Footnotes (if any). These should include a note number or letter as well as text.

   7. Page headers and footers (if any). These are the thin, one-or-two-line headers or footers that tend to appear at the top or bottom of pages, often with a title and page number. (Do not consider larger, more-substantive headers or footers with real page content to be headers and footers, but rather part of the main body text.)

   8. Watermarks or other background images or design elements that are purely decorative and are not needed to understand the meaning of the page content (if any). These you want to ignore completely.

   9. Other (if any). Any other content that doesn't fit into one of the categories above.

Then, respond in correctly-formatted JSON according to the format described below."""

        json_output_spec = f"""Your JSON response should include a single key named `elements` that is a list of dicts, one for each element. Each of these element-specific dicts should include the keys listed below. These elements should be ordered as a human reader is meant to read them (generally from left to right and top to bottom, but might vary depending on the visual layout of the page or pages).

   1. `type` (string): This must be one of the following values, according to the element type descriptions above.

      a. "body_text_section": One section of main body text (or all of the main body text on the page when there are no clear section breaks). Don't forget to capture the page or section heading, if any, which might be stylized in some way.
      b. "boxout": One section of boxout text
      c. "table": One table
      d. "chart": One chart or graph
      e. "image": One image or figure (but remember that watermarks, backgrounds, and purely-decorative images should be ignored and not considered page elements in your JSON output)
      f. "footnote": One footnote
      g. "page_header": One page header
      h. "page_footer": One page footer
      i. "other": One other element that doesn't fit into one of the above categories

   2. `content` (string): This is the content of the element, in markdown format. The content should depend on the `type` as follows:

      a. "body_text_section": A markdown version of the text in the section, beginning with the section header (if any). The text should be verbatim, with no omissions, additions, or revisions other than to format the text in appropriate markdown syntax and remove soft hyphens that were added at the ends of lines when possible (and when the words are not naturally hyphenated). Do not add hyperlink formatting to the markdown.
      b. "boxout": A markdown version of the text in the section, beginning with the section header (if any). The text should be verbatim, with no omissions, additions, or revisions other than to format the text in appropriate markdown syntax and remove soft hyphens that were added at the ends of lines when possible (and when the words are not naturally hyphenated). Do not add hyperlink formatting to the markdown.
      c. "table": A markdown version of the complete table, including any title, column and row labels, cell text or data, and any end notes.
      d. "chart": Describe the chart or graph as if to a blind person, using markdown text and exact details and numbers whenever possible. Be sure to include any title, labels, notes, captions, or other information presented with the chart or graph (generally just above, below, or to the side of the chart or graph). Approximate numeric values from the visual elements, using the axes, in order to report the approximate numeric scale of features in the graph or chart.
      e. "image": Describe the image or figure as if to a blind person, using markdown text and exact details and numbers whenever possible. Be sure to include any title, labels, notes, captions, or other information presented with the image or figure (generally just above, below, or to the side of the image or figure). Do not add hyperlink formatting to the markdown.
      f. "footnote": Markdown text with the exact footnote, including the number or label identifying the footnote.
      g. "page_header": Markdown text with the exact header.
      h. "page_footer": Markdown text with the exact footer.
      i. "other": Markdown text with the content of the element.

Be sure to follow these JSON instructions faithfully, returning a single `elements` object list (each with `type` and `content` keys)."""

        # process PDF to JSON
        all_dicts = self.pdf_to_json(pdf_path, json_context, json_job, json_output_spec)

        # aggregate all elements into a single list
        all_elements = []
        for d in all_dicts:
            all_elements.extend(d['elements'])

        # drop all headers and footers, re-order body text to flow continuously within sections
        all_elements = self._clean_and_reorder_elements(all_elements)

        # assemble and return markdown output
        all_markdown = self._assemble_markdown(all_elements)
        return all_markdown


class ExcelDocumentConverter:
    """Utility class to convert Excel files Markdown tables (if they don't have any images or charts)."""

    class ExcelContent:
        """Class for representing Excel file content."""

        @dataclass
        class TableRange:
            """Represents a contiguous table range in a worksheet."""

            start_row: int
            end_row: int
            start_col: int
            end_col: int
            has_header: bool = True
            is_pivot_table: bool = False

        # initialize class-level member variables
        wb: Workbook = None
        filepath: str = None

        def __init__(self, filepath: str):
            """
            Initialize the Excel content object.

            :param filepath: Path to the Excel file.
            :type filepath: str
            """

            self.wb = load_workbook(filepath, data_only=True)
            self.filepath = filepath

        def has_unsupported_content(self) -> Tuple[bool, str]:
            """
            Check if workbook contains content that we don't support.
            Only checks for images and charts, allowing all other formatting to be quietly lost
            in conversion.

            :return: Tuple indicating if PDF conversion is needed and the reason why.
            :rtype: Tuple[bool, str]
            """

            for sheet in self.wb.worksheets:
                # check for images
                # noinspection PyProtectedMember
                if hasattr(sheet, '_images') and sheet._images:
                    return True, f"Sheet '{sheet.title}' contains images"

                # check for charts
                # noinspection PyProtectedMember
                if hasattr(sheet, '_charts') and sheet._charts:
                    return True, f"Sheet '{sheet.title}' contains charts"

            return False, "Content is suitable for direct markdown conversion"

        @staticmethod
        def _is_pivot_table(sheet: Worksheet, table_range: TableRange) -> bool:
            """
            Detect if a table range is likely a pivot table based on characteristics.
            This function uses a combination of heuristics to detect pivot tables in Excel.

            :param sheet: Worksheet object containing the table range.
            :type sheet: Worksheet
            :param table_range: TableRange object representing the table range.
            :type table_range: TableRange
            :return: True if the table range is likely a pivot table, False otherwise.
            :rtype: bool
            """

            # check if sheet has pivot tables defined
            if hasattr(sheet, 'pivotTables') and sheet.pivotTables:
                # check if our range intersects with any pivot table range
                for pivot in sheet.pivotTables:
                    if (pivot.location.min_row <= table_range.end_row and
                            pivot.location.max_row >= table_range.start_row and
                            pivot.location.min_col <= table_range.end_col and
                            pivot.location.max_col >= table_range.start_col):
                        return True

            # use fallback detection based on pivot table characteristics
            # noinspection PyBroadException
            try:
                first_row = list(sheet.iter_rows(
                    min_row=table_range.start_row,
                    max_row=table_range.start_row,
                    min_col=table_range.start_col,
                    max_col=table_range.end_col
                ))[0]

                # check for common pivot table indicators
                pivot_indicators = ['Total', 'Grand Total', 'Sum of', 'Count of', 'Average of']
                cell_values = [str(cell.value).strip() if cell.value else '' for cell in first_row]
                if any(any(indicator in value for indicator in pivot_indicators)
                       for value in cell_values):
                    return True
            except:
                pass

            return False

        @staticmethod
        def find_tables(sheet: Worksheet) -> List[TableRange]:
            """
            Identify contiguous table ranges in a worksheet.

            :param sheet: Worksheet object to analyze.
            :type sheet: Worksheet
            :return: List of TableRange objects representing the identified table ranges.
            :rtype: List[TableRange]
            """

            tables = []
            current_table = None

            # first check for explicitly defined tables
            if hasattr(sheet, 'tables'):
                for table in sheet.tables.values():
                    table_range = ExcelDocumentConverter.ExcelContent.TableRange(
                        start_row=table.ref.min_row,
                        end_row=table.ref.max_row,
                        start_col=table.ref.min_col,
                        end_col=table.ref.max_col,
                        has_header=True  # (Excel tables always have headers)
                    )
                    tables.append(table_range)

            # Then look for implicit tables in the data
            data_rows = list(sheet.rows)
            if not data_rows:
                return tables

            for row_idx, row in enumerate(data_rows, 1):
                row_empty = all(cell.value is None for cell in row)
                row_merged = any(isinstance(cell, MergedCell) for cell in row)

                if not row_empty and not row_merged:
                    # find the start and end columns for this row
                    start_col = None
                    end_col = None
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value is not None:
                            if start_col is None:
                                start_col = col_idx
                            end_col = col_idx

                    if start_col is not None:
                        if current_table is None:
                            # start new table
                            current_table = ExcelDocumentConverter.ExcelContent.TableRange(
                                start_row=row_idx,
                                end_row=row_idx,
                                start_col=start_col,
                                end_col=end_col
                            )
                        else:
                            # extend current table
                            current_table.end_row = row_idx
                            current_table.start_col = min(current_table.start_col, start_col)
                            current_table.end_col = max(current_table.end_col, end_col)
                else:
                    # empty row - close current table if it exists
                    if current_table is not None:
                        # check if it overlaps with any existing tables
                        if not any(ExcelDocumentConverter.ExcelContent._ranges_overlap(current_table, t) for t in tables):
                            tables.append(current_table)
                        current_table = None

            # add final table if exists and doesn't overlap
            if (current_table is not None and
                    not any(ExcelDocumentConverter.ExcelContent._ranges_overlap(current_table, t) for t in tables)):
                tables.append(current_table)

            # analyze tables for headers and pivot tables
            for table in tables:
                table.has_header = ExcelDocumentConverter.ExcelContent._detect_header_row(sheet, table)
                table.is_pivot_table = ExcelDocumentConverter.ExcelContent._is_pivot_table(sheet, table)

            # return identified tables
            return tables

        @staticmethod
        def _ranges_overlap(range1: TableRange, range2: TableRange) -> bool:
            """
            Check if two table ranges overlap.

            :param range1: First TableRange object.
            :type range1: TableRange
            :param range2: Second TableRange object.
            :type range2: TableRange
            :return: True if the two ranges overlap, False otherwise.
            :rtype: bool
            """

            return not (range1.end_row < range2.start_row or
                        range1.start_row > range2.end_row or
                        range1.end_col < range2.start_col or
                        range1.start_col > range2.end_col)

        @staticmethod
        def _detect_header_row(sheet: Worksheet, table: TableRange) -> bool:
            """
            Detect if the first row of a table range is likely a header row.

            :param sheet: Worksheet object containing the table range.
            :type sheet: Worksheet
            :param table: TableRange object representing the table range.
            :type table: TableRange
            :return: True if the first row is likely a header row, False otherwise.
            :rtype: bool
            """

            if table.start_row == table.end_row:
                return False

            first_row = list(sheet.iter_rows(
                min_row=table.start_row,
                max_row=table.start_row,
                min_col=table.start_col,
                max_col=table.end_col
            ))[0]

            second_row = list(sheet.iter_rows(
                min_row=table.start_row + 1,
                max_row=table.start_row + 1,
                min_col=table.start_col,
                max_col=table.end_col
            ))[0]

            # check for header formatting
            header_indicators = 0

            # check for bold font in first row
            if any(cell.font and cell.font.bold for cell in first_row):
                header_indicators += 1

            # check for different font properties between first and second row
            if any(
                    first_row[i].font != second_row[i].font
                    for i in range(len(first_row))
                    if first_row[i].font and second_row[i].font
            ):
                header_indicators += 1

            # check if first row values look like headers
            first_row_values = [str(cell.value).strip() if cell.value else '' for cell in first_row]

            # headers often have different data types than the data
            if any(isinstance(first_row[i].value, str) and
                   isinstance(second_row[i].value, (int, float))
                   for i in range(len(first_row))):
                header_indicators += 1

            # headers typically don't contain empty cells
            if all(val for val in first_row_values):
                header_indicators += 1

            # consider it a header if we have at least 2 indicators
            return header_indicators >= 2

    @staticmethod
    def _excel_to_strftime_format(excel_format: str) -> str:
        """
        Convert an Excel date format to a strftime format.

        :param excel_format: Excel date format.
        :type excel_format: str
        :return: strftime format.
        :rtype: str
        """

        # define a mapping from Excel format codes to strftime format codes
        format_mapping = {
            'yyyy': '%Y',
            'yy': '%y',
            'mmmm': '%B',
            'mmm': '%b',
            'mm': '%m',  # month as a zero-padded decimal number
            'm': '%m',  # month as a zero-padded decimal number
            'dddd': '%A',
            'ddd': '%a',
            'dd': '%d',
            'd': '%d',  # day of the month as a zero-padded decimal number
            'hh': '%H',
            'h': '%H',  # hour (24-hour clock) as a zero-padded decimal number
            'ss': '%S',
            's': '%S',  # second as a zero-padded decimal number
            'AM/PM': '%p',
            'am/pm': '%p'
        }

        # create a regex pattern that matches any of the Excel format codes
        pattern = re.compile('|'.join(re.escape(key) for key in sorted(format_mapping.keys(), key=len, reverse=True)))

        # replace Excel format codes with strftime format codes
        strftime_format = pattern.sub(lambda x: format_mapping[x.group()], excel_format)

        return strftime_format

    @staticmethod
    def _format_cell_value(cell: Cell, value: Any) -> str:
        """
        Format cell value with appropriate markdown styling based on cell format.

        :param cell: Cell object to format.
        :type cell: Cell
        :param value: Value to format.
        :type value: Any
        :return: Formatted value as a string.
        :rtype: str
        """

        # if no value, just return an empty string
        if value is None:
            return ''

        # handle dates and times
        if cell.is_date and value:
            # noinspection PyBroadException
            try:
                from datetime import datetime
                if isinstance(value, datetime):
                    # do our best to format dates and times as specified in the document, otherwise use a default format
                    strftime_format = ExcelDocumentConverter._excel_to_strftime_format(cell.number_format
                                                                                       or '%Y-%m-%d %H:%M:%S')
                    return value.strftime(strftime_format)
                return str(value)
            except:
                return str(value)

        # handle numbers
        if isinstance(value, (int, float)):
            number_format = cell.number_format or 'General'
            if number_format == 'General':
                # for whole numbers, don't show decimal places
                if isinstance(value, int) or value.is_integer():
                    return str(int(value))
                return str(value)
            elif number_format.endswith('%'):
                # handle percentages
                match = re.search(r'0+(\.0+)?%', number_format)
                if match:
                    # if we find requested number of decimal places, respect it
                    decimal_places = match.group(1).count('0') if match.group(1) else 0
                    return f"{value * 100:.{decimal_places}f}%"
                else:
                    # default to 2 decimal places
                    return f"{value * 100:.2f}%"
            elif '#' in number_format or '0' in number_format:
                # try to respect decimal places specified in format (otherwise default to 2 for non-integers)
                decimal_places = 2
                match = re.search(r'0+(\.0+)?', number_format)
                if match:
                    decimal_places = match.group(1).count('0') if match.group(1) else 0
                if decimal_places == 0 or isinstance(value, int) or value.is_integer():
                    return str(int(value))
                return f"{value:.{decimal_places}f}"

            return str(value)

        # handle text by stripping extra spacing and escaping pipe characters
        value = str(value).strip()
        formatted = value.replace('|', '\\|')

        # add hyperlinks if needed
        if cell.hyperlink:
            if cell.hyperlink.target:
                formatted = f'[{formatted}]({cell.hyperlink.target})'

        # add text formatting if needed (note: order matters here!)
        if cell.font:
            if cell.font.strike:
                formatted = f'~~{formatted}~~'          # strikethrough
            if cell.font.bold:
                formatted = f'**{formatted}**'          # bold
            if cell.font.italic:
                formatted = f'*{formatted}*'            # italic
            if cell.font.vertAlign == 'superscript':
                formatted = f'<sup>{formatted}</sup>'   # superscript
            elif cell.font.vertAlign == 'subscript':
                formatted = f'<sub>{formatted}</sub>'   # subscript

        # handle multi-line text if needed (since it's going to go in a Markdown table, can't leave newlines as-is)
        if '\n' in formatted:
            formatted = formatted.replace('\n', '<br>')

        return formatted

    @staticmethod
    def _get_merge_range_for_cell(cell: Cell, merge_ranges: List[MergedCellRange]) -> MergedCellRange | None:
        """
        Get the merge range containing this cell, if any.

        :param cell: Cell object to check.
        :type cell: Cell
        :param merge_ranges: List of merged cell ranges in the worksheet.
        :type merge_ranges: List[MergedCellRange]
        :return: Merged cell range containing the cell, or None if not merged.
        :rtype: MergedCellRange | None
        """

        # can for cell in supplied merge ranges
        for merge_range in merge_ranges:
            if cell.coordinate in merge_range:
                return merge_range
        return None

    @staticmethod
    def _is_first_cell_in_merge_range(cell: Cell, merge_range: MergedCellRange) -> bool:
        """
        Check if cell is the top-left cell in its merge range.

        :param cell: Cell object to check.
        :type cell: Cell
        :param merge_range: MergedCellRange containing the cell.
        :type merge_range: MergedCellRange
        :return: True if cell is the top-left cell in the merge range, False otherwise.
        :rtype: bool
        """

        return cell.row == merge_range.min_row and cell.column == merge_range.min_col

    @staticmethod
    def _create_markdown_table(sheet: Worksheet, table_range: ExcelContent.TableRange) -> str:
        """
        Convert a table range to Markdown format.

        :param sheet: Worksheet object containing the table range.
        :type sheet: Worksheet
        :param table_range: TableRange object representing the table range.
        :type table_range: TableRange
        :return: Markdown-formatted table.
        :rtype: str
        """

        # track merge ranges that affect our table
        relevant_merges = [
            merge_range for merge_range in sheet.merged_cells.ranges
            if (merge_range.min_row <= table_range.end_row and
                merge_range.max_row >= table_range.start_row and
                merge_range.min_col <= table_range.end_col and
                merge_range.max_col >= table_range.start_col)
        ]

        # initialize
        rows = []
        merge_notes = []

        # handle pivot table headers specially
        if table_range.is_pivot_table:
            # process headers differently for pivot tables
            header_rows = []
            data_start_row = table_range.start_row

            # collect all header rows (those with merged cells or different formatting)
            curr_row = table_range.start_row
            while curr_row < table_range.end_row:
                row = list(sheet.iter_rows(
                    min_row=curr_row,
                    max_row=curr_row,
                    min_col=table_range.start_col,
                    max_col=table_range.end_col
                ))[0]

                if any(isinstance(cell, MergedCell) or
                       (cell.font and cell.font.bold) for cell in row):
                    header_rows.append(row)
                    data_start_row = curr_row + 1
                else:
                    break
                curr_row += 1

            # process header rows
            if header_rows:
                for row in header_rows:
                    row_values = []
                    for cell in row:
                        merge_range = ExcelDocumentConverter._get_merge_range_for_cell(cell, relevant_merges)
                        if merge_range:
                            if ExcelDocumentConverter._is_first_cell_in_merge_range(cell, merge_range):
                                value = sheet.cell(merge_range.min_row, merge_range.min_col).value
                                formatted_value = ExcelDocumentConverter._format_cell_value(cell, value)
                                row_values.append(formatted_value)
                            else:
                                row_values.append('')
                        else:
                            value = cell.value
                            formatted_value = ExcelDocumentConverter._format_cell_value(cell,
                                                                                        value) if value is not None else ''
                            row_values.append(formatted_value)
                    rows.append(row_values)

            # update table range to exclude processed headers
            table_range.start_row = data_start_row
            table_range.has_header = False  # Headers already processed

        # get all cells in range
        for row in sheet.iter_rows(
                min_row=table_range.start_row,
                max_row=table_range.end_row,
                min_col=table_range.start_col,
                max_col=table_range.end_col
        ):
            row_values = []

            for cell in row:
                merge_range = ExcelDocumentConverter._get_merge_range_for_cell(cell, relevant_merges)
                if merge_range:
                    if ExcelDocumentConverter._is_first_cell_in_merge_range(cell, merge_range):
                        value = sheet.cell(merge_range.min_row, merge_range.min_col).value
                        formatted_value = ExcelDocumentConverter._format_cell_value(cell, value)
                        row_values.append(formatted_value)

                        # consider adding comment about the merge
                        span_cols = merge_range.max_col - merge_range.min_col + 1
                        span_rows = merge_range.max_row - merge_range.min_row + 1
                        if span_cols > 1 or span_rows > 1:
                            # add comment about merged cell
                            curr_row = len(rows) + 1
                            curr_col = len(row_values)
                            merge_notes.append(
                                f"* Cell at row {curr_row}, column {curr_col} " +
                                f"spans {span_rows} row{'' if span_rows == 1 else 's'} and " +
                                f"{span_cols} column{'' if span_cols == 1 else 's'}"
                            )
                    else:
                        row_values.append('')
                else:
                    value = cell.value
                    formatted_value = ExcelDocumentConverter._format_cell_value(cell, value) if value is not None else ''
                    row_values.append(formatted_value)

            # only add rows that aren't completely empty
            if any(val.strip() for val in row_values):
                rows.append(row_values)

        if not rows:
            return ''

        # create Markdown table
        md_lines = []

        # add headers (either pivot table headers or regular headers)
        header_row = rows[0]
        md_lines.append('| ' + ' | '.join(header_row) + ' |')

        # add separator
        separator = '|'
        for _ in range(len(header_row)):
            separator += ' --- |'
        md_lines.append(separator)

        # add data rows (skip first row only if it's a regular header)
        if table_range.has_header and not table_range.is_pivot_table:
            data_rows = rows[1:]
        else:
            data_rows = rows[1:] if rows else []  # (for pivot tables, first row is already a header)

        for row in data_rows:
            md_lines.append('| ' + ' | '.join(row) + ' |')

        # add merge notes (if any)
        if merge_notes:
            md_lines.extend(['', '<!-- Merged cells:', *merge_notes, '-->'])

        # combine all lines together for output
        return '\n'.join(md_lines)

    @staticmethod
    def convert_excel_to_markdown(excel_path: str, include_hidden_sheets: bool = False,
                                  lose_unsupported_content: bool = False) -> Tuple[bool, str]:
        """
        Convert Excel file to Markdown if possible, otherwise indicate PDF conversion needed.

        :param excel_path: Path to the Excel file.
        :type excel_path: str
        :param include_hidden_sheets: Whether to include hidden sheets in the conversion. Default is False.
        :type include_hidden_sheets: bool
        :param lose_unsupported_content: Whether to quietly lose unsupported content in the conversion (if False, will
          return failure when file contains images and/or charts). Default is False.
        :type lose_unsupported_content: bool
        :return: Tuple indicating if conversion was successful and the Markdown text.
        """

        try:
            excel_content = ExcelDocumentConverter.ExcelContent(excel_path)
            # drop out if we have images or charts and we're not losing them
            if not lose_unsupported_content:
                has_unsupported_content, reason = excel_content.has_unsupported_content()
                if has_unsupported_content:
                    return False, reason

            # process each sheet
            markdown_content = []

            for sheet in excel_content.wb.worksheets:
                # skip hidden sheets if desired
                if sheet.sheet_state == 'hidden' and not include_hidden_sheets:
                    continue

                markdown_content.append(f'# {sheet.title}\n')

                # find and convert tables
                tables = excel_content.find_tables(sheet)
                if not tables:
                    markdown_content.append('*No data found in this sheet*\n')
                    continue

                # group tables by row to maintain original layout
                row_grouped_tables = {}
                for table in tables:
                    row_grouped_tables.setdefault(table.start_row, []).append(table)

                # process tables in order by row
                last_end_row = 0
                for start_row in sorted(row_grouped_tables.keys()):
                    # add spacing if there's a gap between tables
                    if last_end_row > 0 and start_row > last_end_row + 1:
                        markdown_content.append('\n---\n')

                    # process tables that start on this row
                    row_tables = row_grouped_tables[start_row]

                    # sort tables by column if multiple tables on same row
                    row_tables.sort(key=lambda t: t.start_col)

                    for i, table in enumerate(row_tables):
                        if i > 0:
                            # add horizontal separator between side-by-side tables
                            markdown_content.append('\n___\n')

                        # add table title if it's a pivot table
                        if table.is_pivot_table:
                            # try to find a title from cells above the table
                            title = ExcelDocumentConverter._find_table_title(sheet, table)
                            if title:
                                markdown_content.append(f'### {title}\n')

                        # convert table to markdown
                        markdown_table = ExcelDocumentConverter._create_markdown_table(sheet, table)
                        if markdown_table:
                            markdown_content.append(f"{markdown_table}\n")

                        last_end_row = max(last_end_row, table.end_row)

            # return Markdown text
            return True, '\n'.join(markdown_content)

        except Exception as e:
            return False, f"Error processing file: {str(e)}"

    @staticmethod
    def _find_table_title(sheet: Worksheet, table: ExcelContent.TableRange) -> str | None:
        """
        Look for a title above the table by checking for merged cells or
        cells with larger/bold font.

        :param sheet: Worksheet object containing the table.
        :type sheet: Worksheet
        :param table: TableRange object representing the table.
        :type table: TableRange
        :return: Title text if found, None otherwise.
        :rtype: str | None
        """

        if table.start_row <= 1:
            return None

        # look up to 3 rows above the table
        for row in range(table.start_row - 1, max(0, table.start_row - 4), -1):
            # check cells in and around the table's column range
            start_col = max(1, table.start_col - 1)
            end_col = min(sheet.max_column, table.end_col + 1)

            for cell in sheet[row][start_col - 1:end_col]:
                # skip empty cells
                if not cell.value:
                    continue

                # check if cell is merged
                is_merged = isinstance(cell, MergedCell)
                if is_merged:
                    merge_range = next(
                        merge_range for merge_range in sheet.merged_cells.ranges
                        if cell.coordinate in merge_range
                    )
                    if merge_range.min_row == row and merge_range.min_col == cell.column:
                        master_cell = sheet.cell(merge_range.min_row, merge_range.min_col)
                        return str(master_cell.value).strip()

                # check for title-like formatting
                if cell.font:
                    font_size = cell.font.size if hasattr(cell.font, 'size') else None
                    if ((font_size and font_size > 11) or  # larger font
                            cell.font.bold or  # bold
                            (  # centered over table
                                    cell.alignment and
                                    cell.alignment.horizontal == 'center' and
                                    abs((end_col - start_col) / 2 + start_col - cell.column) <= 1
                            )):
                        return str(cell.value).strip()

        return None


class UnstructuredDocumentConverter:
    """Convert various document types to markdown using Unstructured."""

    @dataclass
    class DocumentElement:
        """Represents a processed document element."""

        type: str
        content: str
        metadata: Optional[Dict] = None
        level: Optional[int] = None

    # initialize class-level member variables
    heading_style: str = "atx"
    element_handlers: Dict[type, Callable] = {}

    def __init__(self, heading_style: str = "atx"):
        """
        Initialize the Unstructured document converter.

        :param heading_style: 'atx' for # style or 'setext' for underline style.
        :type heading_style: str
        """

        # set heading style
        self.heading_style = heading_style

        # map Unstructured element types to markdown handlers
        self.element_handlers = {
            Title: UnstructuredDocumentConverter._handle_title,
            Header: UnstructuredDocumentConverter._handle_header,
            Text: UnstructuredDocumentConverter._handle_text,
            NarrativeText: UnstructuredDocumentConverter._handle_narrative,
            ListItem: UnstructuredDocumentConverter._handle_list_item,
            Table: UnstructuredDocumentConverter._handle_table,
            ImageElement: UnstructuredDocumentConverter._handle_image,
            PageBreak: UnstructuredDocumentConverter._handle_page_break,
            Footer: UnstructuredDocumentConverter._handle_footer,
            Address: UnstructuredDocumentConverter._handle_address,
        }

    def convert_to_markdown(self, file_path: Union[str, Path]) -> str:
        """
        Convert document to Markdown format.

        :param file_path: Path to input file.
        :type file_path: Union[str, Path]
        :return: Markdown formatted string.
        :rtype: str
        """

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # partition document using Unstructured
        elements: List[Any] = partition(str(file_path))

        # process elements to DocumentElement objects
        doc_elements = []
        for element in elements:
            handler = self.element_handlers.get(type(element))
            if handler:
                # noinspection PyArgumentList
                processed = handler(element)
                if processed:
                    doc_elements.extend(processed if isinstance(processed, list) else [processed])
            else:
                # default to handling as text
                processed = UnstructuredDocumentConverter.DocumentElement(
                    type="text",
                    content=element.text.strip(),
                    metadata={"coordinates": getattr(element, "coordinates", None)}
                )
                doc_elements.extend([processed])

        # convert to markdown
        return self._elements_to_markdown(doc_elements)

    @staticmethod
    def _handle_title(element: Title) -> DocumentElement:
        """
        Process title elements.

        :param element: Title element to process.
        :type element: Title
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="heading",
            content=element.text.strip(),
            level=1,
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    @staticmethod
    def _handle_header(element: Header) -> DocumentElement:
        """
        Process header elements.

        :param element: Header element to process.
        :type element: Header
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="heading",
            content=element.text.strip(),
            level=2,
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    @staticmethod
    def _handle_text(element: Text) -> DocumentElement:
        """
        Process basic text elements.

        :param element: Text element to process.
        :type element: Text
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="text",
            content=element.text.strip(),
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    @staticmethod
    def _handle_narrative(element: NarrativeText) -> DocumentElement:
        """
        Process narrative text elements.

        :param element: NarrativeText element to process.
        :type element: NarrativeText
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="paragraph",
            content=element.text.strip(),
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    @staticmethod
    def _handle_list_item(element: ListItem) -> DocumentElement:
        """
        Process list items.

        :param element: ListItem element to process.
        :type element: ListItem
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="list_item",
            content=element.text.strip(),
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "indent_level": getattr(element, "indent_level", 0)
            }
        )

    @staticmethod
    def _handle_table(element: Table) -> DocumentElement | None:
        """
        Process table elements.

        :param element: Table element to process.
        :type element: Table
        :return: DocumentElement object or None if table is not supported.
        :rtype: DocumentElement | None
        """

        # extract table text and format as Markdown table
        table_text = element.text.strip()

        # lines are rows
        rows = table_text.split('\n')
        if not rows:
            return None

        # create Markdown table, assuming whitespace separates columns and the first row is the header
        md_table = ['| ' + ' | '.join(rows[0].split()) + ' |', '|' + '---|' * (len(rows[0].split()) - 1) + '---|']
        for row in rows[1:]:
            md_table.append('| ' + ' | '.join(row.split()) + ' |')

        return UnstructuredDocumentConverter.DocumentElement(
            type="table",
            content='\n'.join(md_table),
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    @staticmethod
    def _handle_image(element: ImageElement) -> DocumentElement:
        """
        Process image elements.

        :param element: ImageElement to process.
        :type element: ImageElement
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        # for now, we just add a placeholder for the image
        return UnstructuredDocumentConverter.DocumentElement(
            type="image",
            content="![Image]",
            metadata={
                "coordinates": getattr(element, "coordinates", None),
                "image_data": getattr(element, "image_data", None)
            }
        )

    @staticmethod
    def _handle_page_break(element: PageBreak) -> DocumentElement:
        """
        Process page breaks.

        :param element: PageBreak element to process.
        :type element: PageBreak
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="page_break",
            content="",
            metadata={"page_number": getattr(element, "page_number", None)}
        )

    @staticmethod
    def _handle_footer(element: Footer) -> DocumentElement:
        """
        Process footer elements.

        :param element: Footer element to process.
        :type element: Footer
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="footer",
            content=element.text.strip(),
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    @staticmethod
    def _handle_address(element: Address) -> DocumentElement:
        """
        Process address elements.

        :param element: Address element to process.
        :type element: Address
        :return: DocumentElement object.
        :rtype: DocumentElement
        """

        return UnstructuredDocumentConverter.DocumentElement(
            type="address",
            content=element.text.strip(),
            metadata={"coordinates": getattr(element, "coordinates", None)}
        )

    def _elements_to_markdown(self, elements: List[DocumentElement]) -> str:
        """
        Convert processed elements to markdown string.

        :param elements: List of DocumentElement objects.
        :type elements: List[DocumentElement]
        :return: Markdown formatted string.
        :rtype: str
        """

        # keep track as we process each element
        markdown_parts = []
        list_stack = []
        for element in elements:
            if element.type == "heading":
                # clear any active lists
                list_stack = []

                # add heading
                if self.heading_style == "atx":
                    markdown_parts.append(f"{'#' * element.level} {element.content}")
                else:
                    markdown_parts.append(element.content)
                    markdown_parts.append('=' if element.level == 1 else '-' * len(element.content))
            elif element.type == "paragraph":
                # clear any active lists
                list_stack = []

                # add paragraph content
                markdown_parts.append(element.content)
            elif element.type == "list_item":
                indent_level = element.metadata.get("indent_level", 0)
                # adjust list stack
                while len(list_stack) > indent_level:
                    list_stack.pop()
                while len(list_stack) < indent_level:
                    list_stack.append(0)

                # add list item
                prefix = "  " * indent_level + "* "
                markdown_parts.append(f"{prefix}{element.content}")
            elif element.type == "table":
                # clear any active lists
                list_stack = []

                # add table content
                markdown_parts.append(element.content)
            elif element.type == "image":
                # clear any active lists
                list_stack = []

                # add image content
                markdown_parts.append(element.content)
            elif element.type == "page_break":
                # add page break
                markdown_parts.append("\n---\n")
            elif element.type == "footer":
                # add footer content
                markdown_parts.append(f"\n---\n{element.content}\n")
            elif element.type == "address":
                # add address content
                markdown_parts.append(f"> {element.content}")
            else:
                # add any other content as-is
                markdown_parts.append(element.content)

            # add spacing between elements
            markdown_parts.append("")

        # return combined markdown content
        return "\n".join(markdown_parts)
